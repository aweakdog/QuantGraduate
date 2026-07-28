"""
Export v29 intersection rebalance JSON to Excel, matching the format of
wf_daily_v25_w100_ts2023-01-01_te2026-07-16_cap100000_close.xlsx

Sheets: 操作清单, 每日汇总, 交易明细, 实际成交, 板块统计
"""
import argparse, json
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]


def load_sector_data():
    names_path = ROOT / "data" / "raw" / "all_stock_list.parquet"
    names_df = pd.read_parquet(names_path)
    names = dict(zip(names_df["code"].astype(str), names_df["name"]))
    concept_path = ROOT / "data" / "universe" / "concept_stock_map.json"
    concept_data = json.loads(concept_path.read_text())
    reverse = defaultdict(list)
    for concept, codes in concept_data["concept_to_stocks"].items():
        for code in codes:
            reverse[str(code)[:6]].append(concept)
    return names, reverse


def format_workbook(path: Path):
    book = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 10), 32)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        headers = {cell.value: cell.column for cell in sheet[1]}
        if "实际日收益" in headers and sheet.max_row > 1:
            col = get_column_letter(headers["实际日收益"])
            sheet.conditional_formatting.add(
                f"{col}2:{col}{sheet.max_row}",
                ColorScaleRule(start_type="min", start_color="F4CCCC",
                               mid_type="num", mid_value=0, mid_color="FFFFFF",
                               end_type="max", end_color="D9EAD3"),
            )
    book.save(path)


def export(source: Path, output: Path):
    result = json.loads(source.read_text())
    names, concepts = load_sector_data()
    daily_raw = result["daily"]
    trades_raw = result["trades"]

    # ── Build trades DataFrame ──
    trades = pd.DataFrame(trades_raw)
    if trades.empty:
        print("No trades to export.")
        return

    trades["name"] = trades["code"].map(lambda c: names.get(str(c)[:6], ""))
    trades["sector"] = trades["code"].map(lambda c: ", ".join(concepts[str(c)[:6]][:4]))
    trades["action_cn"] = trades["action"].map({
        "buy": "买入", "sell": "卖出", "force_sell": "卖出",
        "reject_buy": "买入拒绝",
    }).fillna(trades["action"])
    trades["timing"] = "收盘"

    # ── Build operation rows (pair buys with sells) ──
    pending = defaultdict(list)
    operation_rows = []

    # Build asset lookup from daily
    asset_by_date = {d["date"]: d["portfolio_value"] for d in daily_raw}
    cost_by_date = {d["date"]: d.get("buy_cost", 0) + d.get("sell_cost", 0) for d in daily_raw}

    for _, row in trades.sort_values(["date", "action"]).iterrows():
        if row["action"] in ("buy",):
            pending[row["code"]].append(row)
        elif row["action"] in ("sell", "force_sell"):
            queue = pending.get(row["code"], [])
            buy = queue.pop(0) if queue else None
            operation_rows.append({
                "推荐日期": buy["date"] if buy is not None else row["date"],
                "实际买入日期": buy["date"] if buy is not None else "未知",
                "买入时点": "收盘" if buy is not None else "未知",
                "股票代码": row["code"],
                "股票名称": row["name"],
                "板块": row["sector"],
                "买入股数": buy["shares"] if buy is not None else None,
                "买入价格": buy["price"] if buy is not None else None,
                "买入金额": buy["gross"] if buy is not None else None,
                "实际卖出日期": row["date"],
                "卖出时点": "收盘",
                "卖出股数": row["shares"],
                "卖出价格": row["price"],
                "卖出金额": row["gross"],
                "毛收益": round(row["gross"] - buy["gross"], 2) if buy is not None else None,
                "操作结束后总资产": asset_by_date.get(str(row["date"])),
                "当天交易手续费": cost_by_date.get(str(row["date"]), 0),
                "状态": "已完成" if row["action"] == "sell" else "期末清仓",
            })

    # Pending (bought but not sold)
    for code, queue in pending.items():
        for buy in queue:
            operation_rows.append({
                "推荐日期": buy["date"], "实际买入日期": buy["date"], "买入时点": "收盘",
                "股票代码": code, "股票名称": buy["name"], "板块": buy["sector"],
                "买入股数": buy["shares"], "买入价格": buy["price"], "买入金额": buy["gross"],
                "实际卖出日期": "尚未卖出", "卖出时点": "", "卖出股数": None,
                "卖出价格": None, "卖出金额": None, "毛收益": None,
                "操作结束后总资产": asset_by_date.get(str(buy["date"])),
                "当天交易手续费": cost_by_date.get(str(buy["date"]), 0),
                "状态": "持仓未卖出",
            })

    # Rejected
    rejected = trades[trades["action"] == "reject_buy"]
    for _, row in rejected.iterrows():
        operation_rows.append({
            "推荐日期": row["date"], "实际买入日期": row["date"], "买入时点": "收盘",
            "股票代码": row["code"], "股票名称": row["name"], "板块": row["sector"],
            "买入股数": None, "买入价格": row.get("price"), "买入金额": None,
            "实际卖出日期": "未成交", "卖出时点": "", "卖出股数": None,
            "卖出价格": None, "卖出金额": None, "毛收益": None,
            "操作结束后总资产": asset_by_date.get(str(row["date"])),
            "当天交易手续费": cost_by_date.get(str(row["date"]), 0),
            "状态": f"买入拒绝：{row.get('reason', '')}",
        })

    operations = pd.DataFrame(operation_rows)

    # ── Daily summary ──
    daily = pd.DataFrame(daily_raw)
    if not daily.empty:
        daily["holdings"] = daily["holdings"].apply(lambda v: ", ".join(v) if isinstance(v, list) else str(v))
        daily["to_sell"] = daily["to_sell"].apply(lambda v: ", ".join(v) if isinstance(v, list) else str(v))
        daily["to_buy"] = daily["to_buy"].apply(lambda v: ", ".join(v) if isinstance(v, list) else str(v))
        daily["to_keep"] = daily["to_keep"].apply(lambda v: ", ".join(v) if isinstance(v, list) else str(v))

    daily_columns = {
        "date": "推荐日期", "n_train": "训练样本数", "n_test": "股票数量",
        "ic": "信息系数IC", "portfolio_value": "组合净值", "daily_ret": "实际日收益",
        "n_holdings": "持仓数量", "holdings": "当前持仓",
        "to_sell": "卖出", "to_buy": "买入", "to_keep": "继续持有",
        "sell_cost": "卖出手续费", "buy_cost": "买入手续费",
    }
    daily = daily.rename(columns={k: v for k, v in daily_columns.items() if k in daily.columns})

    # ── Trade detail ──
    trade_detail = trades.rename(columns={
        "date": "推荐日期", "action_cn": "操作", "timing": "交易时点",
        "code": "股票代码", "name": "股票名称", "sector": "板块",
        "shares": "股数", "price": "成交价格", "gross": "成交金额",
        "reason": "原因",
    })
    trade_detail["实际交易日期"] = trade_detail["推荐日期"]

    actual = trade_detail[trade_detail["操作"].isin(["买入", "卖出"])].copy()

    # ── Sector summary ──
    sector_rows = []
    for _, row in trades.iterrows():
        for sector in concepts[str(row["code"])[:6]]:
            sector_rows.append({"sector": sector, "action": row["action_cn"], "code": row["code"]})
    sector_df = pd.DataFrame(sector_rows)
    if sector_df.empty:
        sector_summary = pd.DataFrame(columns=["板块", "推荐或交易记录数", "买入次数", "卖出次数"])
    else:
        sector_summary = sector_df.groupby("sector").agg(
            recommended_or_traded_rows=("code", "size"),
            buy_count=("action", lambda v: int((v == "买入").sum())),
            sell_count=("action", lambda v: int((v == "卖出").sum())),
        ).reset_index().sort_values("recommended_or_traded_rows", ascending=False)
        sector_summary.columns = ["板块", "推荐或交易记录数", "买入次数", "卖出次数"]

    # ── Write Excel ──
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        operations.to_excel(writer, sheet_name="操作清单", index=False)
        daily.to_excel(writer, sheet_name="每日汇总", index=False)
        trade_detail.to_excel(writer, sheet_name="交易明细", index=False)
        actual.to_excel(writer, sheet_name="实际成交", index=False)
        sector_summary.to_excel(writer, sheet_name="板块统计", index=False)
    format_workbook(output)
    print(f"created: {output}")
    print(f"daily_rows={len(daily)} trade_rows={len(trade_detail)} actual_rows={len(actual)}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or args.source.with_suffix(".xlsx")
    export(args.source, output)


if __name__ == "__main__":
    main()
