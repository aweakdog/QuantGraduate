"""把 v35 回测结果导出为可执行的操作 Excel

默认输出到桌面: ~/Desktop/量化操作表/<生成日期>/<结果名>.xlsx

用法:
  python scripts/export_v35_excel.py data/processed/wf_daily_*.json
  python scripts/export_v35_excel.py <json> --output /自定义/路径.xlsx
"""
import argparse, json
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DESKTOP_ROOT = Path.home() / "Desktop" / "量化操作表"


def default_output(src: Path) -> Path:
    """桌面/量化操作表/<生成日期>/<结果名>.xlsx"""
    return DESKTOP_ROOT / datetime.now().strftime("%Y-%m-%d") / f"{src.stem}.xlsx"


def load_meta():
    names = {}
    # PIT 元数据覆盖全市场 5563 只, 优先用它兜底; all_stock_list 再覆盖一次
    pm = ROOT / "data" / "universe" / "pit_metadata.parquet"
    if pm.exists():
        m = pd.read_parquet(pm)
        names = dict(zip(m["code"].astype(str).str.zfill(6), m["name"]))
    p = ROOT / "data" / "raw" / "all_stock_list.parquet"
    if p.exists():
        n = pd.read_parquet(p)
        names.update(dict(zip(n["code"].astype(str).str[:6], n["name"])))
    concepts = defaultdict(list)
    cp = ROOT / "data" / "universe" / "concept_stock_map.json"
    if cp.exists():
        cd = json.loads(cp.read_text())
        for concept, codes in cd.get("concept_to_stocks", {}).items():
            for c in codes:
                concepts[str(c)[:6]].append(concept)
    return names, concepts


def benchmark_series(dates, train_file="training_data_v24.parquet", pit_universe=None,
                     skip_boards=()):
    """池内等权买入持有基准的日收益, 与回测口径一致

    开启 PIT 成分约束时, 基准只统计当期生效成分股 (与回测完全同口径)。
    主板-only 回测(skip_boards)的基准同样剪掉受限板块 —— 否则每日汇总表的
    超额收益列和绩效摘要(读回测自带的主板基准)会对不上。
    """
    df = pd.read_parquet(ROOT / "data" / "processed" / train_file,
                         columns=["date", "code", "fwd_1d_ret"])
    df["date"] = pd.to_datetime(df["date"])
    df = df.dropna(subset=["fwd_1d_ret"])
    if skip_boards:
        df = df[~df["code"].astype(str).str.startswith(tuple(skip_boards))]
    if pit_universe:
        u = pd.read_parquet(ROOT / "data" / "universe" / pit_universe)
        u["effective_date"] = pd.to_datetime(u["effective_date"])
        eff = pd.DatetimeIndex(sorted(u["effective_date"].unique()))
        members = {d: set(g["code"].astype(str).str.zfill(6))
                   for d, g in u.groupby("effective_date")}
        c6 = df["code"].astype(str).str[:6]
        per = np.searchsorted(eff, df["date"].values, side="right") - 1
        keep = np.zeros(len(df), bool)
        for i, d in enumerate(eff):
            m = per == i
            if m.any():
                keep[m] = c6[m].isin(members[pd.Timestamp(d)]).values
        df = df[keep]
    b = df.groupby("date")["fwd_1d_ret"].mean().shift(1)
    return b.reindex(pd.to_datetime(dates)).fillna(0.0).values


def build_operations(trades, names, concepts, pv_by_date):
    """FIFO 配对 买入 -> 卖出, 生成一行一笔完整操作"""
    pending = defaultdict(deque)
    rows = []
    for t in sorted(trades, key=lambda x: (x["date"], x["action"] != "sell")):
        code = str(t["code"])[:6]
        if t["action"] == "buy":
            pending[code].append(t)
            continue
        buy = pending[code].popleft() if pending[code] else None
        if buy is None:
            continue
        hold_days = (pd.Timestamp(t["date"]) - pd.Timestamp(buy["date"])).days
        gross_pnl = t["gross"] - buy["gross"]
        net_pnl = t["net"] + buy["net"]        # buy 的 net 已是负数
        rows.append({
            "买入日期": buy["date"], "卖出日期": t["date"], "自然持有天数": hold_days,
            "股票代码": code, "股票名称": names.get(code, ""),
            "板块": ", ".join(concepts.get(code, [])[:3]),
            "股数": buy["shares"], "买入价": round(buy["price"], 3), "卖出价": round(t["price"], 3),
            "买入金额": round(buy["gross"], 2), "卖出金额": round(t["gross"], 2),
            "买入费用": round(buy["fee"], 2), "卖出费用": round(t["fee"], 2),
            "毛收益": round(gross_pnl, 2), "净收益": round(net_pnl, 2),
            "收益率%": round(net_pnl / buy["gross"] * 100, 2) if buy["gross"] else None,
            "卖出原因": {"matured": "持满到期", "end": "回测期末清仓"}.get(t.get("reason"), t.get("reason")),
            "卖出后总资产": pv_by_date.get(t["date"]),
            "状态": "已完成",
        })
    for code, q in pending.items():
        for buy in q:
            rows.append({
                "买入日期": buy["date"], "卖出日期": "尚未卖出", "自然持有天数": None,
                "股票代码": code, "股票名称": names.get(code, ""),
                "板块": ", ".join(concepts.get(code, [])[:3]),
                "股数": buy["shares"], "买入价": round(buy["price"], 3), "卖出价": None,
                "买入金额": round(buy["gross"], 2), "卖出金额": None,
                "买入费用": round(buy["fee"], 2), "卖出费用": None,
                "毛收益": None, "净收益": None, "收益率%": None,
                "卖出原因": "", "卖出后总资产": None, "状态": "持仓中",
            })
    return pd.DataFrame(rows).sort_values(["买入日期", "股票代码"]).reset_index(drop=True)


def format_workbook(path, n_ops):
    book = load_workbook(path)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True)
    for ws in book.worksheets:
        if ws.max_row < 1:
            continue
        ws.freeze_panes = "A2"
        if ws.title != "绩效摘要":
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col[:400]) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w, 10), 34)
        hdr = {c.value: c.column for c in ws[1]}
        for key in ("净收益", "收益率%", "实际日收益", "超额收益", "毛收益"):
            if key in hdr and ws.max_row > 1:
                L = get_column_letter(hdr[key])
                ws.conditional_formatting.add(
                    f"{L}2:{L}{ws.max_row}",
                    ColorScaleRule(start_type="min", start_color="F4CCCC",
                                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                                   end_type="max", end_color="D9EAD3"))
    book.save(path)


def export(src: Path, out: Path):
    res = json.loads(src.read_text())
    names, concepts = load_meta()
    s = res["summary"]

    daily = pd.DataFrame(res["daily"])
    pv_by_date = dict(zip(daily["date"], daily["portfolio_value"]))

    # ── 每日汇总 ──
    bench = benchmark_series(daily["date"], res.get("train_file", "training_data_v24.parquet"),
                            res.get("pit_universe"), tuple(res.get("skip_boards") or ()))
    daily["基准日收益"] = bench
    daily["超额收益"] = daily["daily_ret"] - bench
    daily["策略净值"] = (1 + daily["daily_ret"]).cumprod()
    daily["基准净值"] = (1 + daily["基准日收益"]).cumprod()
    daily["持仓明细"] = daily["holdings"].apply(
        lambda hs: ", ".join(f"{str(c)[:6]}{names.get(str(c)[:6], '')}" for c in hs) if isinstance(hs, list) else "")
    daily["空仓"] = (daily["in_cash"].astype(bool).map({True: "空仓", False: ""})
                     if "in_cash" in daily.columns else "")
    dsum = daily.rename(columns={
        "date": "日期", "portfolio_value": "组合总资产", "cash": "现金",
        "daily_ret": "实际日收益", "n_holdings": "持仓只数", "deployed": "仓位占比",
        "sell_cost": "卖出费用", "buy_cost": "买入费用", "ic": "信息系数IC",
    })[["日期", "空仓", "组合总资产", "现金", "仓位占比", "持仓只数", "实际日收益", "基准日收益",
        "超额收益", "策略净值", "基准净值", "买入费用", "卖出费用", "信息系数IC", "持仓明细"]]
    for c in ("组合总资产", "现金"):
        dsum[c] = dsum[c].round(2)
    for c in ("策略净值", "基准净值"):
        dsum[c] = dsum[c].round(4)

    # ── 操作清单 ──
    ops = build_operations(res["trades"], names, concepts, pv_by_date)

    # ── 交易明细 ──
    tr = pd.DataFrame(res["trades"])
    tr["股票名称"] = tr["code"].map(lambda c: names.get(str(c)[:6], ""))
    tr["板块"] = tr["code"].map(lambda c: ", ".join(concepts.get(str(c)[:6], [])[:3]))
    tr["action"] = tr["action"].map({"buy": "买入", "sell": "卖出", "force_sell": "期末清仓"}).fillna(tr["action"])
    tr["reason"] = tr["reason"].map({"new_tranche": "新开一档", "matured": "持满到期",
                                     "end": "回测期末", "regime_exit": "大盘转弱清仓"}).fillna(tr["reason"])
    tr = tr.rename(columns={"date": "日期", "code": "股票代码", "action": "操作", "shares": "股数",
                            "price": "成交价", "gross": "成交金额", "fee": "手续费",
                            "net": "现金变动", "reason": "原因"})
    tr = tr[["日期", "操作", "股票代码", "股票名称", "板块", "股数", "成交价",
             "成交金额", "手续费", "现金变动", "原因"]]

    # ── 绩效摘要 ──
    done = ops[ops["状态"] == "已完成"]
    win = (done["净收益"] > 0).mean() * 100 if len(done) else 0
    meta = [
        ("回测区间", res["period"]),
        ("交易日数", res["n_days"]),
        ("初始资金", f"¥{res['initial_capital']:,.0f}"),
        ("期末资产", f"¥{s['final_value']:,.2f}"),
        ("标签(预测目标)", res["label"]),
        ("中性化方式", "仅按日期demean（不做概念中性化）"),
        ("持有天数/分档数", res["hold_days"]),
        ("每档买入只数", res["tranche_n"]),
        ("目标持仓只数", res["target_positions"]),
        ("入选特征数", res["features"]),
        ("训练集", res.get("train_file", "training_data_v24.parquet")),
        ("股票池约束", res.get("pit_universe") or "无(全期通用池)"),
        ("——— 交易成本 ———", ""),
        ("佣金费率", f"{res.get('trade_cost', 0.0006)*100:.3f}% /边 "
                     f"(最低¥{res.get('min_fee', 5):.0f})"),
        ("滑点", f"{res.get('slippage', 0)*100:.3f}% /边"),
        ("往返总成本", f"{(res.get('trade_cost', 0.0006)+res.get('slippage', 0))*2*100:.3f}%"),
        ("——— 空仓择时 ———", ""),
        ("空仓判据", res.get("regime_filter", "off")),
        ("广度阈值", res.get("regime_breadth")),
        ("趋势均线窗口", res.get("regime_ma")),
        ("切换确认天数", res.get("regime_confirm")),
        ("空仓天数", f"{s.get('cash_days', 0)} 天 ({s.get('cash_days_pct', 0):.1f}%)"),
        ("——— 收益 ———", ""),
        ("总收益率", f"{s['total_return_pct']:+.1f}%"),
        ("年化收益率", f"{s['annualized_return_pct']:+.1f}%"),
        ("夏普比率", s["sharpe"]),
        ("最大回撤", f"{s['max_dd_pct']:.1f}%"),
        ("——— 对比等权买入持有基准 ———", ""),
        ("基准总收益", f"{s['benchmark_total_pct']:+.1f}%"),
        ("基准年化", f"{s['benchmark_annual_pct']:+.1f}%"),
        ("年化超额收益", f"{s['excess_annual_pct']:+.1f}%"),
        ("信息比率IR", s["information_ratio"]),
        ("beta", s["beta"]),
        ("年化alpha(剔除beta)", f"{s['alpha_annual_pct']:+.1f}%"),
        ("是否跑赢基准", "是" if s["beat_benchmark"] else "否"),
        ("——— 信号质量 ———", ""),
        ("IC均值", s["ic_mean"]),
        ("IC标准差", s["ic_std"]),
        ("IC t统计量", f"{s['ic_tstat']}（>2 才显著）"),
        ("——— 执行 ———", ""),
        ("平均仓位占比", f"{s['avg_deployed_pct']:.1f}%"),
        ("平均持仓只数", s["avg_holdings"]),
        ("总手续费占初始资金", f"{s['total_cost_pct']:.1f}%"),
        ("成交笔数", s["n_trades"]),
        ("完整操作笔数", len(done)),
        ("单笔胜率", f"{win:.1f}%"),
        ("单笔平均收益率", f"{done['收益率%'].mean():.2f}%" if len(done) else "-"),
        ("买入拒单次数", s["rejected_buy"]),
        ("卖出拒单次数", s["rejected_sell"]),
    ]
    summary = pd.DataFrame(meta, columns=["指标", "数值"])

    # ── 板块统计 ──
    if len(done):
        rows = []
        for _, r in done.iterrows():
            for sec in concepts.get(r["股票代码"], []) or ["未分类"]:
                rows.append({"板块": sec, "净收益": r["净收益"], "收益率%": r["收益率%"]})
        sec = (pd.DataFrame(rows).groupby("板块")
               .agg(操作笔数=("净收益", "size"), 净收益合计=("净收益", "sum"),
                    平均收益率=("收益率%", "mean"), 胜率=("净收益", lambda v: (v > 0).mean() * 100))
               .reset_index().sort_values("净收益合计", ascending=False))
        for c in ("净收益合计", "平均收益率", "胜率"):
            sec[c] = sec[c].round(2)
    else:
        sec = pd.DataFrame(columns=["板块", "操作笔数", "净收益合计", "平均收益率", "胜率"])

    # ── 个股统计 ──
    if len(done):
        stk = (done.groupby(["股票代码", "股票名称"])
               .agg(操作笔数=("净收益", "size"), 净收益合计=("净收益", "sum"),
                    平均收益率=("收益率%", "mean"), 胜率=("净收益", lambda v: (v > 0).mean() * 100))
               .reset_index().sort_values("净收益合计", ascending=False))
        for c in ("净收益合计", "平均收益率", "胜率"):
            stk[c] = stk[c].round(2)
    else:
        stk = pd.DataFrame()

    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        summary.to_excel(w, sheet_name="绩效摘要", index=False)
        ops.to_excel(w, sheet_name="操作清单", index=False)
        dsum.to_excel(w, sheet_name="每日汇总", index=False)
        tr.to_excel(w, sheet_name="交易明细", index=False)
        stk.to_excel(w, sheet_name="个股统计", index=False)
        sec.to_excel(w, sheet_name="板块统计", index=False)
    format_workbook(out, len(ops))
    print(f"已生成: {out}")
    print(f"  回测区间 {res['period']} | 操作 {len(ops)} 笔 (已完成 {len(done)}) | 成交 {len(tr)} 笔")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("source", type=Path)
    ap.add_argument("--output", type=Path, default=None)
    a = ap.parse_args()
    export(a.source, a.output or default_output(a.source))
