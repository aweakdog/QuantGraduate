"""把回测结果 json 导出成一个 Excel 工作簿, 方便逐日/逐笔核对

工作表:
  1. 汇总对比   —— 各配置关键指标并排, 含基准
  2. 分年度     —— 每个配置逐年 策略/基准/超额
  3. 净值曲线   —— 逐日净值 + 基准净值 + 空仓标记, 附折线图
  4. 每日明细   —— 单个配置逐日 持仓数/现金/成本/IC/持仓代码
  5. 交易明细   —— 逐笔 买卖/股数/价格/费用, 带筛选
  6. 入选特征   —— 模型实际使用的特征清单

用法:
  python scripts/export_backtest_excel.py                       # 默认导出 PIT 池全部配置
  python scripts/export_backtest_excel.py --pattern "data/processed/wf_daily_pit_*.json"
  python scripts/export_backtest_excel.py --detail breadth      # 明细表取哪个配置
"""
import argparse
import glob
import json
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
ORDER = {"off": 0, "ma": 1, "breadth": 2, "both": 3, "any": 4}

METRIC_ROWS = [
    ("总收益 %", "total_return_pct"), ("年化收益 %", "annualized_return_pct"),
    ("夏普", "sharpe"), ("最大回撤 %", "max_dd_pct"),
    ("基准总收益 %", "benchmark_total_pct"), ("基准年化 %", "benchmark_annual_pct"),
    ("超额年化 %", "excess_annual_pct"), ("信息比率 IR", "information_ratio"),
    ("beta", "beta"), ("年化 alpha %", "alpha_annual_pct"),
    ("IC 均值", "ic_mean"), ("IC t值", "ic_tstat"),
    ("空仓天数占比 %", "cash_days_pct"), ("总费用 %", "total_cost_pct"),
    ("交易笔数", "n_trades"), ("期末资产", "final_value"),
]


def load_runs(pattern):
    runs = []
    for f in sorted(glob.glob(pattern)):
        d = json.load(open(f, encoding="utf-8"))
        runs.append((d.get("regime_filter", Path(f).stem), d))
    runs.sort(key=lambda x: ORDER.get(x[0], 9))
    return runs


def benchmark_series(train_file, pit_universe, dates):
    """复算池内等权基准日收益 (与回测内部一致: T日实现T-1日标签)"""
    cols = ["date", "code", "fwd_1d_ret"]
    df = pd.read_parquet(ROOT / "data/processed" / train_file, columns=cols)
    df["date"] = pd.to_datetime(df["date"])
    df = df.dropna(subset=["fwd_1d_ret"])
    if pit_universe:
        u = pd.read_parquet(ROOT / "data/universe" / pit_universe)
        u["effective_date"] = pd.to_datetime(u["effective_date"])
        eff = np.array(sorted(u["effective_date"].unique()))
        members = {d: set(g["code"].astype(str).str.zfill(6))
                   for d, g in u.groupby("effective_date")}
        c6 = df["code"].astype(str).str[:6]
        per = np.searchsorted(eff, df["date"].values, side="right") - 1
        keep = np.zeros(len(df), bool)
        for i, d in enumerate(eff):
            m = per == i
            if m.any():
                keep[m] = c6[m].isin(members[pd.Timestamp(d)]).values
        df = df[keep]
    b = df.groupby("date")["fwd_1d_ret"].mean().shift(1)
    return b.reindex(pd.DatetimeIndex(dates)).fillna(0.0)


def sheet_summary(runs, xw):
    rows = []
    for label, mkey in METRIC_ROWS:
        row = {"指标": label}
        for name, d in runs:
            row[name] = d["summary"].get(mkey)
        rows.append(row)
    cfg = [("空仓判据", "regime_filter"), ("广度阈值", "regime_breadth"),
           ("确认天数", "regime_confirm"), ("持有天数", "hold_days"),
           ("目标持仓数", "target_positions"), ("执行方式", "exec_mode"),
           ("初始资金", "initial_capital"), ("训练集", "train_file"),
           ("成分约束", "pit_universe"), ("特征数", "features"), ("回测区间", "period")]
    rows.append({"指标": ""})
    for label, k in cfg:
        rows.append({"指标": label, **{name: d.get(k) for name, d in runs}})
    pd.DataFrame(rows).to_excel(xw, sheet_name="汇总对比", index=False)


def sheet_yearly(runs, bench_map, xw):
    out = []
    for name, d in runs:
        dd = pd.DataFrame(d["daily"])
        dd["date"] = pd.to_datetime(dd["date"])
        b = bench_map[name]
        dd["bench"] = b.values
        for year, g in dd.groupby(dd["date"].dt.year):
            s = (1 + g["daily_ret"]).prod() - 1
            bm = (1 + g["bench"]).prod() - 1
            out.append({"配置": name, "年份": year, "交易日": len(g),
                        "策略 %": round(s * 100, 1), "基准 %": round(bm * 100, 1),
                        "超额 %": round((s - bm) * 100, 1),
                        "空仓天数": int(g["in_cash"].sum()) if "in_cash" in g else 0,
                        "胜": "✓" if s > bm else "✗"})
    pd.DataFrame(out).to_excel(xw, sheet_name="分年度", index=False)


def sheet_equity(runs, bench_map, xw):
    eq = None
    for name, d in runs:
        dd = pd.DataFrame(d["daily"])
        dd["date"] = pd.to_datetime(dd["date"])
        cap = d["initial_capital"]
        cur = pd.DataFrame({
            "日期": dd["date"],
            f"{name}_净值": (dd["portfolio_value"] / cap).round(4),
            f"{name}_空仓": dd["in_cash"].astype(int) if "in_cash" in dd else 0,
        })
        eq = cur if eq is None else eq.merge(cur, on="日期", how="outer")
    name0 = runs[0][0]
    eq["基准_净值"] = (1 + bench_map[name0].values).cumprod().round(4)
    eq.to_excel(xw, sheet_name="净值曲线", index=False)


def sheet_detail(run, xw):
    name, d = run
    dd = pd.DataFrame(d["daily"])
    dd["holdings"] = dd["holdings"].apply(
        lambda x: ",".join(map(str, x)) if isinstance(x, list) else x)
    ren = {"date": "日期", "portfolio_value": "总资产", "cash": "现金",
           "daily_ret": "日收益", "n_holdings": "持仓数", "holdings": "持仓代码",
           "deployed": "已投入", "sell_cost": "卖出成本", "buy_cost": "买入成本",
           "ic": "当日IC", "in_cash": "空仓"}
    dd.rename(columns=ren).to_excel(xw, sheet_name=f"每日明细_{name}", index=False)

    tr = pd.DataFrame(d["trades"])
    ren2 = {"date": "成交日", "signal_date": "信号日", "code": "代码", "action": "方向",
            "shares": "股数", "price": "价格", "gross": "金额", "fee": "费用",
            "net": "净额", "reason": "原因"}
    tr.rename(columns=ren2).to_excel(xw, sheet_name=f"交易明细_{name}", index=False)


def sheet_features(runs, xw):
    feats = runs[0][1].get("selected_features", [])

    def cat(f):
        if f.startswith(("mf_", "dde_", "mtss_", "fund_flow")):
            return "资金流"
        if f.startswith(("ev_", "tev_")):
            return "事件"
        if f.startswith(("ann_", "has_ann", "days_since_ann")):
            return "公告"
        if f.startswith(("pe", "pb", "roe", "eps", "bps", "revenue", "profit",
                         "debt_", "gross_", "total_assets")):
            return "基本面"
        if f.startswith(("cn_", "us", "a50", "sp_", "dj_", "nq_", "sox", "mkt_", "usd")):
            return "宏观商品"
        if f.startswith("con_"):
            return "概念板块"
        return "技术面"

    pd.DataFrame({"特征": feats, "类别": [cat(f) for f in feats]}) \
        .to_excel(xw, sheet_name="入选特征", index=False)


def polish(path):
    """加图表 + 冻结首行 + 列宽 + 筛选"""
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col[:200] if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 9), 42)
        if ws.max_row > 2 and ws.title not in ("汇总对比",):
            ws.auto_filter.ref = ws.dimensions

    ws = wb["净值曲线"]
    ch = LineChart()
    ch.title = "净值曲线 (起点=1.0)"
    ch.height, ch.width = 11, 26
    ch.y_axis.title = "净值"
    val_cols = [i for i, c in enumerate(ws[1], 1)
                if c.value and ("净值" in str(c.value))]
    for ci in val_cols:
        ch.add_data(Reference(ws, min_col=ci, min_row=1, max_row=ws.max_row), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
    ws.add_chart(ch, f"{chr(65 + len(ws[1]) + 1)}2")
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pattern", default="data/processed/wf_daily_pit_*_cap100000.json")
    ap.add_argument("--detail", default=None, help="明细表用哪个配置, 默认最优夏普")
    ap.add_argument("--out", default=None)
    a = ap.parse_args()

    runs = load_runs(a.pattern)
    if not runs:
        raise SystemExit(f"没找到结果文件: {a.pattern}")
    print(f"读取 {len(runs)} 个配置: {[n for n, _ in runs]}")

    bench_map = {}
    for name, d in runs:
        dates = pd.to_datetime([x["date"] for x in d["daily"]])
        bench_map[name] = benchmark_series(d["train_file"], d.get("pit_universe"), dates)
        print(f"  {name}: {len(dates)} 个交易日, 基准复算完成")

    if a.detail:
        run = next(r for r in runs if r[0] == a.detail)
    else:
        run = max(runs, key=lambda r: r[1]["summary"]["sharpe"])
    print(f"明细表配置: {run[0]}")

    out = Path(a.out) if a.out else ROOT / "data/processed/backtest_report.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        sheet_summary(runs, xw)
        sheet_yearly(runs, bench_map, xw)
        sheet_equity(runs, bench_map, xw)
        sheet_detail(run, xw)
        sheet_features(runs, xw)
    polish(out)
    print(f"已导出: {out}")


if __name__ == "__main__":
    main()
