import argparse
import json
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]


def load_sector_data():
    names_path = ROOT / "data" / "raw" / "all_stock_list.parquet"
    names_df = pd.read_parquet(names_path)
    names = dict(zip(names_df["code"].astype(str), names_df["name"]))
    concept_path = ROOT / "data" / "universe" / "concept_stock_map.json"
    concept_data = json.loads(concept_path.read_text())
    reverse = defaultdict(list)
    for concept, codes in concept_data["concept_to_stocks"].items():
        for code in codes:
            reverse[str(code)[:6]].append(concept)
    return names, reverse


def format_workbook(path: Path):
    book = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 10), 32)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        headers = {cell.value: cell.column for cell in sheet[1]}
        if "daily_ret" in headers and sheet.max_row > 1:
            col = get_column_letter(headers["daily_ret"])
            sheet.conditional_formatting.add(
                f"{col}2:{col}{sheet.max_row}",
                ColorScaleRule(start_type="min", start_color="F4CCCC", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="D9EAD3"),
            )
    book.save(path)


def export(source: Path, output: Path):
    result = json.loads(source.read_text())
    names, concepts = load_sector_data()
    daily = pd.DataFrame(result["daily"])
    portfolio = pd.DataFrame(result["portfolio"]["daily_values"])
    trades = pd.DataFrame(result.get("execution", {}).get("trades", []))

    if not daily.empty:
        daily["trade_date"] = pd.to_datetime(daily["date"]) + pd.offsets.BDay(1)
        daily["holdings"] = daily["holdings"].apply(lambda values: ", ".join(values))
        daily["rand_holdings"] = daily["rand_holdings"].apply(lambda values: ", ".join(values))
    asset_by_signal_date = {}
    if not portfolio.empty:
        portfolio = portfolio.rename(columns={"value": "portfolio_value"})
        asset_by_signal_date = dict(zip(portfolio["date"].astype(str), portfolio["portfolio_value"]))
        daily = daily.merge(portfolio[["date", "portfolio_value", "daily_ret", "turnover_bought", "turnover_sold", "buy_cost", "sell_cost", "total_cost"]], on="date", how="left")
    asset_by_trade_date = {}
    cost_by_trade_date = {}
    if not portfolio.empty:
        for signal_date, group in trades.groupby("signal_date"):
            trade_dates = group["trade_date"].dropna().astype(str)
            if len(trade_dates) > 0 and str(signal_date) in asset_by_signal_date:
                trade_date = trade_dates.iloc[0]
                asset_by_trade_date[trade_date] = asset_by_signal_date[str(signal_date)]
                portfolio_row = portfolio[portfolio["date"].astype(str) == str(signal_date)]
                if not portfolio_row.empty:
                    cost_by_trade_date[trade_date] = portfolio_row.iloc[0].get("total_cost", 0)

    if not trades.empty:
        trades["name"] = trades["code"].map(names).fillna("")
        trades["sector"] = trades["code"].map(lambda code: ", ".join(concepts[str(code)[:6]][:4]))
        trades["gross_amount"] = trades["gross_amount"].fillna(0)
        trades["action"] = trades["action"].map({"BUY": "买入", "SELL": "卖出", "BUY_REJECTED": "买入拒绝", "SELL_REJECTED": "卖出拒绝"}).fillna(trades["action"])
        trades["timing"] = trades["timing"].map({"open": "开盘", "close": "收盘"}).fillna(trades["timing"])
        trades["reason"] = trades["reason"].map({
            "integer_lot": "整数手成交",
            "daily_exit": "当日收盘退出",
            "insufficient_one_lot": "本金不足一手",
            "limit_up_open": "开盘涨停无法买入",
            "limit_down_open": "开盘跌停无法卖出",
            "limit_down_close": "收盘跌停无法卖出",
            "missing_bar": "缺少行情",
            "carry_position": "前日未卖出持仓",
        }).fillna(trades["reason"])
        actual = trades[trades["action"].isin(["买入", "卖出"])].copy()
    else:
        actual = trades.copy()

    pending = defaultdict(list)
    operation_rows = []
    for _, row in trades.sort_values(["trade_date", "timing"]).iterrows():
        if row["action"] == "买入":
            pending[row["code"]].append(row)
        elif row["action"] == "卖出":
            queue = pending.get(row["code"], [])
            buy = queue.pop(0) if queue else None
            operation_rows.append({
                "推荐日期": buy["signal_date"] if buy is not None else row["signal_date"],
                "实际买入日期": buy["trade_date"] if buy is not None else "未知",
                "买入时点": buy["timing"] if buy is not None else "未知",
                "股票代码": row["code"],
                "股票名称": row["name"],
                "板块": row["sector"],
                "买入股数": buy["shares"] if buy is not None else None,
                "买入价格": buy["price"] if buy is not None else None,
                "买入金额": buy["gross_amount"] if buy is not None else None,
                "实际卖出日期": row["trade_date"],
                "卖出时点": row["timing"],
                "卖出股数": row["shares"],
                "卖出价格": row["price"],
                "卖出金额": row["gross_amount"],
                "毛收益": round(row["gross_amount"] - buy["gross_amount"], 2) if buy is not None else None,
                "操作结束后总资产": asset_by_trade_date.get(str(row["trade_date"])),
                "当天交易手续费": cost_by_trade_date.get(str(row["trade_date"]), 0),
                "状态": "已完成",
            })

    for code, queue in pending.items():
        for buy in queue:
            operation_rows.append({
                "推荐日期": buy["signal_date"], "实际买入日期": buy["trade_date"], "买入时点": buy["timing"],
                "股票代码": code, "股票名称": buy["name"], "板块": buy["sector"],
                "买入股数": buy["shares"], "买入价格": buy["price"], "买入金额": buy["gross_amount"],
                "实际卖出日期": "尚未卖出", "卖出时点": "", "卖出股数": None, "卖出价格": None,
                "卖出金额": None, "毛收益": None,
                "操作结束后总资产": asset_by_trade_date.get(str(buy["trade_date"])),
                "当天交易手续费": cost_by_trade_date.get(str(buy["trade_date"]), 0), "状态": "持仓未卖出",
            })

    rejected = trades[trades["action"].isin(["买入拒绝", "卖出拒绝"])].copy()
    for _, row in rejected.iterrows():
        operation_rows.append({
            "推荐日期": row["signal_date"], "实际买入日期": row["trade_date"] or "未知", "买入时点": row["timing"],
            "股票代码": row["code"], "股票名称": row["name"], "板块": row["sector"],
            "买入股数": None, "买入价格": row.get("price"), "买入金额": None,
            "实际卖出日期": "未成交", "卖出时点": "", "卖出股数": None, "卖出价格": None,
            "卖出金额": None, "毛收益": None,
            "操作结束后总资产": asset_by_trade_date.get(str(row["trade_date"])),
            "当天交易手续费": cost_by_trade_date.get(str(row["trade_date"]), 0),
            "状态": row["action"] + "：" + row["reason"],
        })
    operations = pd.DataFrame(operation_rows)

    sector_rows = []
    for _, row in trades.iterrows():
        for sector in concepts[str(row["code"])[:6]]:
            sector_rows.append({"sector": sector, "action": row["action"], "code": row["code"]})
    sector_df = pd.DataFrame(sector_rows)
    if sector_df.empty:
        sector_summary = pd.DataFrame(columns=["板块", "推荐或交易记录数", "买入次数", "卖出次数"])
    else:
        sector_summary = sector_df.groupby("sector").agg(
            recommended_or_traded_rows=("code", "size"),
            buy_count=("action", lambda values: int((values == "买入").sum())),
            sell_count=("action", lambda values: int((values == "卖出").sum())),
        ).reset_index().sort_values("recommended_or_traded_rows", ascending=False)
        sector_summary.columns = ["板块", "推荐或交易记录数", "买入次数", "卖出次数"]

    daily_columns = {
        "date": "推荐日期", "trade_date": "实际交易日期", "n_train": "训练样本数", "n_test": "股票数量",
        "ic": "信息系数IC", "top3_ret": "Top3超额收益", "top3_raw_ret": "Top3理论原始收益",
        "rand_ret": "随机组合超额收益", "rand_raw_ret": "随机组合原始收益", "hit_rate": "命中率",
        "long_short": "预测均值", "holdings": "模型推荐股票", "rand_holdings": "随机组合股票",
        "portfolio_value": "组合净值", "daily_ret": "实际日收益", "turnover_bought": "实际买入数量", "turnover_sold": "实际卖出数量",
        "buy_cost": "买入手续费", "sell_cost": "卖出手续费", "total_cost": "当天交易手续费",
    }
    daily = daily.rename(columns=daily_columns)
    trade_columns = {
        "signal_date": "推荐日期", "trade_date": "实际交易日期", "action": "操作", "timing": "交易时点",
        "code": "股票代码", "name": "股票名称", "sector": "板块", "shares": "股数", "price": "成交价格",
        "gross_amount": "成交金额", "reason": "原因",
    }
    trades = trades.rename(columns=trade_columns)
    actual = actual.rename(columns=trade_columns)

    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        operations.to_excel(writer, sheet_name="操作清单", index=False)
        daily.to_excel(writer, sheet_name="每日汇总", index=False)
        trades.to_excel(writer, sheet_name="交易明细", index=False)
        actual.to_excel(writer, sheet_name="实际成交", index=False)
        sector_summary.to_excel(writer, sheet_name="板块统计", index=False)
    format_workbook(output)
    print(f"created: {output}")
    print(f"daily_rows={len(daily)} trade_rows={len(trades)} actual_rows={len(actual)}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or args.source.with_suffix(".xlsx")
    export(args.source, output)


if __name__ == "__main__":
    main()
